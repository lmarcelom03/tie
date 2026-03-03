import os
from datetime import date, datetime, timedelta
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from db import (
    init_db,
    add_scheduled_records,
    get_month_records,
    update_records_status_and_notes,
    admin_update_scheduled_date,
    admin_delete_record,
    upsert_records_from_excel,
)

st.set_page_config(page_title="Registro de Actividades", layout="wide")
init_db()

def month_bounds(any_day: date):
    first = any_day.replace(day=1)
    # next month
    if first.month == 12:
        next_first = date(first.year + 1, 1, 1)
    else:
        next_first = date(first.year, first.month + 1, 1)
    last = next_first - timedelta(days=1)
    return first, last

def add_months(base: date, months: int) -> date:
    total = (base.month - 1) + months
    year = base.year + (total // 12)
    month = (total % 12) + 1
    first_of_month = date(year, month, 1)
    if month == 12:
        next_first = date(year + 1, 1, 1)
    else:
        next_first = date(year, month + 1, 1)
    last_day = (next_first - timedelta(days=1)).day
    return first_of_month.replace(day=min(base.day, last_day))

def get_admin_code() -> str | None:
    # Prefer Streamlit secrets, fallback to env var
    try:
        return st.secrets.get("ADMIN_CODE")
    except Exception:
        return None

def verify_admin(code: str) -> bool:
    secret = get_admin_code() or os.getenv("ADMIN_CODE")
    return bool(secret) and code == secret

# --- Sidebar / access ---
st.sidebar.title("Acceso")
role = st.sidebar.radio("Rol", ["Especialista", "Administrador"], index=0)
actor_name = st.sidebar.text_input("Tu nombre", value=st.session_state.get("actor_name", ""), placeholder="Ej. Ana Pérez")

is_admin = False
if role == "Administrador":
    code = st.sidebar.text_input("Código especial (admin)", type="password")
    is_admin = verify_admin(code.strip()) if code else False
    st.sidebar.caption("Consejo: guarda el código en `.streamlit/secrets.toml` o variable de entorno `ADMIN_CODE`.")
    if not is_admin and code:
        st.sidebar.error("Código incorrecto.")
else:
    is_admin = False

if actor_name:
    st.session_state["actor_name"] = actor_name.strip()
else:
    st.sidebar.warning("Escribe tu nombre para registrar/actualizar actividades.")

st.sidebar.divider()
today = date.today()
selected_month = st.sidebar.date_input("Mes de trabajo", value=today, help="Se usa para filtrar el tablero y la matriz mensual.")
month_first, month_last = month_bounds(selected_month)

st.title("📋 Registro de Actividades")
st.caption("Registra actividades por día y marca el resultado: ✓ cumplido / ✗ incumplido.")

tab_reg, tab_estado, tab_cal, tab_export, tab_admin = st.tabs(
    ["➕ Registrar", "✅/✗ Mis actividades", "🗓️ Calendario", "⬇️ Exportar", "🛡️ Admin"]
)

# --- Registrar ---
with tab_reg:
    st.subheader("Registrar nueva actividad (programación)")
    with st.form("form_registro", clear_on_submit=True):
        col1, col2, col3 = st.columns([2, 3, 2])
        with col1:
            if role == "Especialista":
                especialista = st.text_input("Especialista", value=actor_name, placeholder="Nombre completo", disabled=True)
            else:
                especialista = st.text_input("Especialista", value=actor_name, placeholder="Nombre completo")
        with col2:
            actividad = st.text_input("Actividad (tarea)", placeholder="Ej. Elaborar informe mensual")
        with col3:
            unidad = st.selectbox(
                "Unidad de medida",
                ["Documento", "Informe", "Reunión", "Visita", "Análisis", "Otro"],
                index=1,
            )
            unidad_otro = st.text_input("Si elegiste 'Otro', especifica", placeholder="Ej. Presentación")
        st.markdown("**Programación**")
        mode = st.radio("Tipo", ["Fecha única", "Rango"], horizontal=True)

        es_rutinaria = st.checkbox("Actividad rutinaria (se repite)", value=False)
        frecuencia = None
        repetir_hasta = None
        if es_rutinaria:
            cfr1, cfr2 = st.columns([2, 2])
            with cfr1:
                frecuencia = st.selectbox("Frecuencia", ["Diaria", "Semanal", "Mensual", "Trimestral", "Semestral"])
            with cfr2:
                repetir_hasta = st.date_input("Repetir hasta", value=month_last)

        if mode == "Fecha única":
            fecha = st.date_input("Fecha programada", value=today)
            fechas = [fecha]
        else:
            c1, c2, c3 = st.columns([1, 1, 2])
            with c1:
                f_ini = st.date_input("Inicio", value=today)
            with c2:
                f_fin = st.date_input("Fin", value=today)
            with c3:
                dias = st.multiselect(
                    "Días de la semana (opcional)",
                    ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"],
                    default=["Lun", "Mar", "Mié", "Jue", "Vie"],
                    help="Si lo dejas como está, se programará de lunes a viernes.",
                )
                incluir_finde = st.checkbox("Incluir fines de semana", value=False)

            # Build date list
            fechas = []
            if f_ini <= f_fin:
                cur = f_ini
                map_wd = {0: "Lun", 1: "Mar", 2: "Mié", 3: "Jue", 4: "Vie", 5: "Sáb", 6: "Dom"}
                allowed = set(dias) if dias else set(map_wd.values())
                weekend_selected = bool(allowed.intersection({"Sáb", "Dom"}))
                while cur <= f_fin:
                    wd = map_wd[cur.weekday()]
                    if wd in allowed and (incluir_finde or wd not in {"Sáb", "Dom"} or weekend_selected):
                        fechas.append(cur)
                    cur += timedelta(days=1)

        notas = st.text_area("Notas / detalle (opcional)", placeholder="Criterios, entregables, etc.")
        submitted = st.form_submit_button("Registrar")

    if submitted:
        if es_rutinaria and fechas:
            inicio = min(fechas)
            if repetir_hasta is None or repetir_hasta < inicio:
                st.error("La fecha 'Repetir hasta' debe ser mayor o igual a la fecha de inicio.")
                fechas = []
            else:
                fechas_rep = []
                cur = inicio
                while cur <= repetir_hasta:
                    fechas_rep.append(cur)
                    if frecuencia == "Diaria":
                        cur += timedelta(days=1)
                    elif frecuencia == "Semanal":
                        cur += timedelta(days=7)
                    elif frecuencia == "Mensual":
                        cur = add_months(cur, 1)
                    elif frecuencia == "Trimestral":
                        cur = add_months(cur, 3)
                    elif frecuencia == "Semestral":
                        cur = add_months(cur, 6)
                    else:
                        break
                fechas = sorted(set(fechas_rep))

        if mode == "Rango" and len(fechas) == 0:
            st.error("No hay fechas válidas en el rango. Revisa inicio/fin y los días seleccionados.")
        elif not (especialista and actividad and (unidad or unidad_otro) and fechas):
            st.error("Faltan datos: especialista, actividad, unidad y al menos 1 fecha.")
        else:
            unidad_final = unidad_otro.strip() if unidad == "Otro" else unidad
            records = [
                {
                    "specialist": especialista.strip(),
                    "activity": actividad.strip(),
                    "unit": unidad_final,
                    "scheduled_date": f.isoformat(),
                    "status": "",
                    "notes": notas.strip(),
                    "created_by": (actor_name or "—").strip(),
                }
                for f in fechas
            ]
            add_scheduled_records(records)
            st.success(f"Listo: se registraron {len(records)} actividad(es).")

# --- Mis actividades del día / Marcar estado ---
with tab_estado:
    st.subheader("Mis actividades del día")
    if not actor_name:
        st.info("Escribe tu nombre en la barra lateral para continuar.")
    else:
        if role == "Especialista":
            rango_ini = today
            rango_fin = today
            df = get_month_records(rango_ini, rango_fin, specialist=actor_name.strip())
            st.caption("Solo se muestran tus actividades programadas para hoy.")
        else:
            c1, c2, c3 = st.columns([2, 2, 2])
            with c1:
                filtro_especialista = st.text_input("Filtrar por especialista", value="")
            with c2:
                rango_ini = st.date_input("Desde", value=month_first)
            with c3:
                rango_fin = st.date_input("Hasta", value=month_last)
            df = get_month_records(rango_ini, rango_fin, specialist=(filtro_especialista.strip() or None))

        if df.empty:
            st.warning("No hay registros en el rango seleccionado.")
        else:
            df = df.sort_values(["scheduled_date", "specialist", "activity"]).reset_index(drop=True)
            original = df.copy()

            st.caption("Edita **Estado** y **Notas**. Solo administrador puede cambiar fecha programada.")
            edited = st.data_editor(
                df,
                hide_index=True,
                use_container_width=True,
                column_config={
                    "id": st.column_config.NumberColumn("ID", disabled=True),
                    "scheduled_date": st.column_config.DateColumn("Fecha", disabled=not is_admin),
                    "specialist": st.column_config.TextColumn("Especialista", disabled=True),
                    "activity": st.column_config.TextColumn("Actividad", disabled=True),
                    "unit": st.column_config.TextColumn("UM", disabled=True),
                    "status": st.column_config.SelectboxColumn("Estado", options=["", "✓", "✗"], required=False),
                    "notes": st.column_config.TextColumn("Notas", width="large"),
                    "updated_at": st.column_config.TextColumn("Actualizado", disabled=True),
                },
            )

            col_save, col_hint = st.columns([1, 3])
            with col_save:
                if st.button("Guardar cambios"):
                    changes = []
                    for _, row in edited.iterrows():
                        rid = int(row["id"])
                        old = original.loc[original["id"] == rid].iloc[0]
                        if not is_admin and row["scheduled_date"] != old["scheduled_date"]:
                            st.error(f"No autorizado: el registro ID {rid} cambió fecha. Se ignorará ese cambio.")
                            row["scheduled_date"] = old["scheduled_date"]
                        if (row["status"] != old["status"]) or (str(row["notes"]) != str(old["notes"])):
                            changes.append(
                                {
                                    "id": rid,
                                    "status": row["status"],
                                    "notes": row["notes"] if pd.notna(row["notes"]) else "",
                                    "actor": actor_name.strip(),
                                }
                            )

                    if not changes:
                        st.info("No detecté cambios para guardar.")
                    else:
                        update_records_status_and_notes(changes)
                        st.success(f"Guardado: {len(changes)} cambio(s).")
                        st.rerun()
            with col_hint:
                st.caption("Tip: deja Estado en blanco si aún no se ejecuta.")

# --- Calendario / recordatorios ---
with tab_cal:
    st.subheader("Calendario de actividades y recordatorios")
    if not actor_name and role == "Especialista":
        st.info("Escribe tu nombre para visualizar tu calendario.")
    else:
        specialist_scope = actor_name.strip() if role == "Especialista" else None
        dfc = get_month_records(month_first, month_last, specialist=specialist_scope)

        if dfc.empty:
            st.info("Aún no hay actividades registradas en el mes seleccionado.")
        else:
            dfc["scheduled_date"] = pd.to_datetime(dfc["scheduled_date"]).dt.date
            dfc["day"] = pd.to_datetime(dfc["scheduled_date"]).dt.day

            if role == "Especialista":
                st.caption("Solo ves tus actividades del mes seleccionado.")
                matriz = (
                    dfc.assign(carga=1)
                    .pivot_table(index="activity", columns="day", values="carga", aggfunc="sum", fill_value=0)
                    .reindex(columns=list(range(1, month_last.day + 1)), fill_value=0)
                    .reset_index()
                )
                st.dataframe(matriz, use_container_width=True, hide_index=True)

                st.markdown("**Próximos deadlines (recordatorio)**")
                proximas = dfc[dfc["scheduled_date"] >= today].sort_values("scheduled_date")[["scheduled_date", "activity", "unit", "status", "notes"]]
                if proximas.empty:
                    st.success("No tienes actividades futuras pendientes en este mes.")
                else:
                    st.dataframe(proximas, use_container_width=True, hide_index=True)
            else:
                st.caption("Vista consolidada por especialista (administración).")
                matriz_carga = (
                    dfc.assign(carga=1)
                    .pivot_table(index="specialist", columns="day", values="carga", aggfunc="sum", fill_value=0)
                    .reindex(columns=list(range(1, month_last.day + 1)), fill_value=0)
                    .reset_index()
                )
                matriz_carga["Total mes"] = matriz_carga.drop(columns=["specialist"]).sum(axis=1)
                st.dataframe(matriz_carga, use_container_width=True, hide_index=True)

# --- Exportar ---
with tab_export:
    st.subheader("Exportar matriz mensual a Excel")
    st.caption("Genera una matriz tipo Excel (días del mes como columnas) con símbolos: ✓ cumplido, ✗ fuera de plazo y + pendiente.")

    st.markdown("### Plantilla de carga masiva")
    st.caption("Descarga esta plantilla, llénala y súbela para actualizar automáticamente registros.")

    plantilla_df = pd.DataFrame([
        {
            "id": "",
            "especialista": "",
            "actividad": "",
            "unidad": "",
            "fecha_programada": "YYYY-MM-DD",
            "estado": "",
            "notas": "",
        }
    ])
    template_bio = BytesIO()
    with pd.ExcelWriter(template_bio, engine="openpyxl") as writer:
        plantilla_df.to_excel(writer, sheet_name="plantilla", index=False)
    st.download_button(
        label="Descargar plantilla para importar",
        data=template_bio.getvalue(),
        file_name="plantilla_carga_actividades.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    archivo_subido = st.file_uploader(
        "Subir Excel para actualización automática",
        type=["xlsx"],
        help="Columnas obligatorias: especialista, actividad, unidad, fecha_programada. Opcionales: id, estado, notas.",
    )

    if archivo_subido is not None:
        try:
            df_in = pd.read_excel(archivo_subido)
            df_in.columns = [str(c).strip().lower() for c in df_in.columns]

            rename_map = {
                "especialista": "specialist",
                "actividad": "activity",
                "unidad": "unit",
                "fecha_programada": "scheduled_date",
                "estado": "status",
                "notas": "notes",
            }
            df_in = df_in.rename(columns=rename_map)

            req = {"specialist", "activity", "unit", "scheduled_date"}
            if not req.issubset(set(df_in.columns)):
                st.error("El archivo no tiene las columnas mínimas requeridas: especialista, actividad, unidad, fecha_programada.")
            else:
                # Normalización básica
                for col in ["id", "status", "notes"]:
                    if col not in df_in.columns:
                        df_in[col] = ""

                if pd.api.types.is_datetime64_any_dtype(df_in["scheduled_date"]):
                    df_in["scheduled_date"] = df_in["scheduled_date"].dt.date.astype(str)
                else:
                    df_in["scheduled_date"] = df_in["scheduled_date"].astype(str).str.strip()

                df_in["status"] = (
                    df_in["status"].astype(str).str.strip().replace({
                        "Cumplido": "✓",
                        "cumplido": "✓",
                        "Incumplido": "✗",
                        "incumplido": "✗",
                        "Pendiente": "",
                        "pendiente": "",
                        "nan": "",
                    })
                )

                records = df_in[["id", "specialist", "activity", "unit", "scheduled_date", "status", "notes"]].fillna("").to_dict("records")

                if st.button("Procesar importación"):
                    actor = (actor_name or "IMPORTADOR").strip()
                    inserted, updated, errors = upsert_records_from_excel(records, actor)
                    st.success(f"Importación completada. Insertados: {inserted} | Actualizados: {updated}")
                    if errors:
                        st.warning("Se detectaron filas con error:")
                        for err in errors[:30]:
                            st.write(f"- {err}")
        except Exception as e:
            st.error(f"No se pudo leer el archivo Excel: {e}")

    st.divider()
    st.markdown("### Exportar matriz")
    if role == "Especialista":
        exp_especialista = actor_name.strip()
        st.caption(f"Exportación limitada a tus registros: {exp_especialista}")
    else:
        exp_especialista = st.text_input("Filtrar por especialista (opcional)", value="")

    if st.button("Generar Excel"):
        df_export = get_month_records(month_first, month_last, specialist=(exp_especialista.strip() or None))

        if df_export.empty:
            wb = Workbook()
            ws = wb.active
            ws.title = "Matriz"
            ws["A1"] = "Sin datos para el rango seleccionado"
            bio = BytesIO()
            wb.save(bio)
            bytes_xlsx = bio.getvalue()
        else:
            # Símbolos para exportación:
            # ✓ = cumplido, ✗ = incumplido (incluye vencidos sin marcar), + = pendiente
            today_ref = date.today()

            def excel_symbol(row) -> str:
                if row["status"] == "✓":
                    return "✓"
                if row["status"] == "✗" or row["scheduled_date"] < today_ref:
                    return "✗"
                return "+"

            df_export["symbol"] = df_export.apply(excel_symbol, axis=1)
            df_export["day"] = pd.to_datetime(df_export["scheduled_date"]).dt.day
            pivot = df_export.pivot_table(
                index=["specialist", "activity", "unit"],
                columns="day",
                values="symbol",
                aggfunc="first",
                fill_value="",
            ).reset_index()

            wb = Workbook()
            ws = wb.active
            ws.title = "Matriz"
            ws.sheet_view.showGridLines = False

            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(color="FFFFFF", bold=True)
            thin = Side(style="thin", color="9E9E9E")
            border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left = Alignment(horizontal="left", vertical="center", wrap_text=True)

            ws["A1"] = f"Matriz mensual: {month_first.strftime('%Y-%m')}"
            ws["A1"].font = Font(size=14, bold=True, color="1F4E79")
            ws.merge_cells("A1:AI1")

            # Leyenda de simbología en el propio Excel
            ws["A2"] = "Leyenda: ✓ cumplido | ✗ incumplido/vencido | + pendiente"
            ws["A2"].font = Font(size=11, color="1F4E79")
            ws.merge_cells("A2:AI2")

            header_row = 4
            headers = ["Especialista", "Actividad", "Unidad de medida"]
            max_day = month_last.day

            for j, h in enumerate(headers, start=1):
                c = ws.cell(row=header_row, column=j, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = center
                c.border = border_thin

            for d in range(1, max_day + 1):
                col = 3 + d
                c = ws.cell(row=header_row, column=col, value=d)
                c.fill = header_fill
                c.font = header_font
                c.alignment = center
                c.border = border_thin

            ws.column_dimensions["A"].width = 24
            ws.column_dimensions["B"].width = 38
            ws.column_dimensions["C"].width = 18
            for d in range(1, max_day + 1):
                ws.column_dimensions[get_column_letter(3 + d)].width = 4.2

            start_row = header_row + 1
            for i, (_, row) in enumerate(pivot.iterrows(), start=start_row):
                ws.cell(i, 1, row["specialist"]).alignment = left
                ws.cell(i, 2, row["activity"]).alignment = left
                ws.cell(i, 3, row["unit"]).alignment = center
                for c in range(1, 3 + max_day + 1):
                    ws.cell(i, c).border = border_thin

                for d in range(1, max_day + 1):
                    val = row[d] if d in row.index else ""
                    ws.cell(i, 3 + d, val).alignment = center

            bio = BytesIO()
            wb.save(bio)
            bytes_xlsx = bio.getvalue()

        filename = f"matriz_{month_first.strftime('%Y_%m')}.xlsx"
        st.download_button(
            label="Descargar Excel",
            data=bytes_xlsx,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# --- Admin ---
with tab_admin:
    if not is_admin:
        st.info("Sección solo para Administrador. Ingresa el código en la barra lateral.")
    else:
        st.subheader("Administración (editar plazos / borrar registros)")
        st.warning("Los cambios quedan registrados en un log de auditoría.")

        c1, c2, c3 = st.columns([1, 2, 2])
        with c1:
            rid = st.number_input("ID del registro", min_value=1, step=1)
        with c2:
            nueva_fecha = st.date_input("Nueva fecha programada", value=today)
        with c3:
            motivo = st.text_input("Motivo del cambio", placeholder="Ej. Reprogramación por solicitud del jefe")

        col_a, col_b = st.columns(2)
        with col_a:
            if st.button("Actualizar fecha (plazo)"):
                ok = admin_update_scheduled_date(
                    record_id=int(rid),
                    new_date=nueva_fecha.isoformat(),
                    actor=actor_name.strip() or "ADMIN",
                    reason=motivo.strip() or "—",
                )
                if ok:
                    st.success("Fecha actualizada.")
                else:
                    st.error("No se pudo actualizar (ID no encontrado).")

        with col_b:
            if st.button("Borrar registro"):
                ok = admin_delete_record(int(rid), actor=actor_name.strip() or "ADMIN", reason=motivo.strip() or "—")
                if ok:
                    st.success("Registro borrado.")
                else:
                    st.error("No se pudo borrar (ID no encontrado).")
