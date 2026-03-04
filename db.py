from __future__ import annotations

import os
import shutil
import sqlite3
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_DIR = Path(os.getenv("APP_DATA_DIR", Path(__file__).resolve().parent / "data"))
DB_PATH = DB_DIR / "app.db"
BACKUP_DIR = DB_DIR / "backups"

def get_db_path() -> str:
    return str(DB_PATH)

def _restore_from_latest_backup_if_needed() -> None:
    if DB_PATH.exists():
        return
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backups = sorted(BACKUP_DIR.glob("app_*.db"), key=lambda p: p.stat().st_mtime, reverse=True)
    if backups:
        DB_PATH.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(backups[0], DB_PATH)

def _backup_db() -> None:
    if not DB_PATH.exists():
        return
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup_name = f"app_{datetime.now().strftime('%Y%m%d')}.db"
    backup_path = BACKUP_DIR / backup_name
    if not backup_path.exists():
        shutil.copy2(DB_PATH, backup_path)

def _now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat(sep=" ")

def get_conn() -> sqlite3.Connection:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    _restore_from_latest_backup_if_needed()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db() -> None:
    with get_conn() as conn:
        conn.execute("""
        CREATE TABLE IF NOT EXISTS scheduled_activities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            specialist TEXT NOT NULL,
            activity TEXT NOT NULL,
            unit TEXT NOT NULL,
            scheduled_date TEXT NOT NULL, -- ISO yyyy-mm-dd
            status TEXT NOT NULL DEFAULT '',
            notes TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            created_by TEXT NOT NULL,
            updated_at TEXT,
            updated_by TEXT
        )
        """)
        conn.execute("""
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action TEXT NOT NULL,
            record_id INTEGER NOT NULL,
            old_scheduled_date TEXT,
            new_scheduled_date TEXT,
            actor TEXT NOT NULL,
            reason TEXT NOT NULL,
            ts TEXT NOT NULL
        )
        """)
        conn.commit()
    _backup_db()

def add_scheduled_records(records: list[dict[str, Any]]) -> None:
    if not records:
        return
    with get_conn() as conn:
        for r in records:
            conn.execute("""
            INSERT INTO scheduled_activities
            (specialist, activity, unit, scheduled_date, status, notes, created_at, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                r["specialist"],
                r["activity"],
                r["unit"],
                r["scheduled_date"],
                r.get("status", "") or "",
                r.get("notes", "") or "",
                _now_iso(),
                r.get("created_by", "—") or "—",
            ))
        conn.commit()
    _backup_db()

def get_month_records(date_from: date, date_to: date, specialist: str | None = None) -> pd.DataFrame:
    q = """
    SELECT id, specialist, activity, unit, scheduled_date, status, notes, updated_at
    FROM scheduled_activities
    WHERE date(scheduled_date) >= date(?) AND date(scheduled_date) <= date(?)
    """
    params: list[Any] = [date_from.isoformat(), date_to.isoformat()]
    if specialist:
        q += " AND specialist = ?"
        params.append(specialist)
    q += " ORDER BY date(scheduled_date) ASC, specialist ASC, activity ASC"
    with get_conn() as conn:
        rows = conn.execute(q, params).fetchall()
    if not rows:
        return pd.DataFrame(columns=["id","scheduled_date","specialist","activity","unit","status","notes","updated_at"])
    df = pd.DataFrame([dict(r) for r in rows])
    df["scheduled_date"] = pd.to_datetime(df["scheduled_date"]).dt.date
    return df

def update_records_status_and_notes(changes: list[dict[str, Any]]) -> None:
    if not changes:
        return
    with get_conn() as conn:
        for ch in changes:
            conn.execute("""
            UPDATE scheduled_activities
            SET status = ?, notes = ?, updated_at = ?, updated_by = ?
            WHERE id = ?
            """, (
                ch.get("status", "") or "",
                ch.get("notes", "") or "",
                _now_iso(),
                ch.get("actor", "—") or "—",
                int(ch["id"]),
            ))
        conn.commit()
    _backup_db()


def upsert_records_from_excel(records: list[dict[str, Any]], actor: str) -> tuple[int, int, list[str]]:
    inserted = 0
    updated = 0
    errors: list[str] = []
    if not records:
        return inserted, updated, errors

    with get_conn() as conn:
        for idx, r in enumerate(records, start=1):
            try:
                specialist = (r.get("specialist") or "").strip()
                activity = (r.get("activity") or "").strip()
                unit = (r.get("unit") or "").strip()
                scheduled_date = (r.get("scheduled_date") or "").strip()
                status = (r.get("status") or "").strip()
                notes = (r.get("notes") or "").strip()
                record_id = r.get("id")

                if not (specialist and activity and unit and scheduled_date):
                    errors.append(f"Fila {idx}: faltan columnas obligatorias (specialist, activity, unit, scheduled_date).")
                    continue
                if status not in {"", "✓", "✗"}:
                    errors.append(f"Fila {idx}: status inválido '{status}'. Usa '', '✓' o '✗'.")
                    continue
                try:
                    date.fromisoformat(scheduled_date)
                except ValueError:
                    errors.append(f"Fila {idx}: scheduled_date inválida '{scheduled_date}'. Usa YYYY-MM-DD.")
                    continue

                if record_id is not None and str(record_id).strip() != "":
                    rec_id = int(record_id)
                    exists = conn.execute("SELECT id FROM scheduled_activities WHERE id = ?", (rec_id,)).fetchone()
                    if exists:
                        conn.execute(
                            """
                            UPDATE scheduled_activities
                            SET specialist = ?, activity = ?, unit = ?, scheduled_date = ?, status = ?, notes = ?, updated_at = ?, updated_by = ?
                            WHERE id = ?
                            """,
                            (specialist, activity, unit, scheduled_date, status, notes, _now_iso(), actor, rec_id),
                        )
                        updated += 1
                        continue

                conn.execute(
                    """
                    INSERT INTO scheduled_activities
                    (specialist, activity, unit, scheduled_date, status, notes, created_at, created_by, updated_at, updated_by)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (specialist, activity, unit, scheduled_date, status, notes, _now_iso(), actor, _now_iso(), actor),
                )
                inserted += 1
            except Exception as e:
                errors.append(f"Fila {idx}: error inesperado ({e}).")
        conn.commit()
    _backup_db()

    return inserted, updated, errors

def admin_update_scheduled_date(record_id: int, new_date: str, actor: str, reason: str) -> bool:
    with get_conn() as conn:
        row = conn.execute("SELECT scheduled_date FROM scheduled_activities WHERE id = ?", (record_id,)).fetchone()
        if not row:
            return False
        old = row["scheduled_date"]
        conn.execute("UPDATE scheduled_activities SET scheduled_date = ?, updated_at = ?, updated_by = ? WHERE id = ?",
                     (new_date, _now_iso(), actor, record_id))
        conn.execute("""
        INSERT INTO audit_log (action, record_id, old_scheduled_date, new_scheduled_date, actor, reason, ts)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """, ("UPDATE_DATE", record_id, old, new_date, actor, reason, _now_iso()))
        conn.commit()
        _backup_db()
        return True

def admin_delete_record(record_id: int, actor: str, reason: str) -> bool:
    with get_conn() as conn:
        row = conn.execute("SELECT scheduled_date FROM scheduled_activities WHERE id = ?", (record_id,)).fetchone()
        if not row:
            return False
        old = row["scheduled_date"]
        conn.execute("DELETE FROM scheduled_activities WHERE id = ?", (record_id,))
        conn.execute("""
        INSERT INTO audit_log (action, record_id, old_scheduled_date, new_scheduled_date, actor, reason, ts)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """, ("DELETE", record_id, old, None, actor, reason, _now_iso()))
        conn.commit()
        _backup_db()
        return True

# ---------- Excel export (matrix style) ----------
def export_month_matrix_xlsx_bytes(month_first: date, month_last: date, specialist: str | None = None) -> bytes:
    df = get_month_records(month_first, month_last, specialist=specialist)
    # Pivot: rows = (specialist, activity, unit), cols = day, values = status
    if df.empty:
        wb = Workbook()
        ws = wb.active
        ws.title = "Matriz"
        ws["A1"] = "Sin datos para el rango seleccionado"
        from io import BytesIO
        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    df["day"] = pd.to_datetime(df["scheduled_date"]).dt.day
    pivot = df.pivot_table(
        index=["specialist", "activity", "unit"],
        columns="day",
        values="status",
        aggfunc="first",
        fill_value="",
    ).reset_index()

    wb = Workbook()
    ws = wb.active
    ws.title = "Matriz"
    ws.sheet_view.showGridLines = False

    # Styles
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="9E9E9E")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws["A1"] = f"Matriz mensual: {month_first.strftime('%Y-%m')}"
    ws["A1"].font = Font(size=14, bold=True, color="1F4E79")
    ws.merge_cells("A1:AI1")

    header_row = 3
    headers = ["Especialista", "Actividad", "Unidad de medida"]
    max_day = month_last.day

    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=j, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border_thin

    # Day columns
    for d in range(1, max_day + 1):
        col = 3 + d
        c = ws.cell(row=header_row, column=col, value=d)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border_thin

    # Column widths
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 18
    for d in range(1, max_day + 1):
        ws.column_dimensions[get_column_letter(3 + d)].width = 4.2

    # Data rows
    start_row = header_row + 1
    for i, row in enumerate(pivot.itertuples(index=False), start=start_row):
        ws.cell(i, 1, row.specialist).alignment = left
        ws.cell(i, 2, row.activity).alignment = left
        ws.cell(i, 3, row.unit).alignment = center
        for c in range(1, 3 + max_day + 1):
            ws.cell(i, c).border = border_thin

        # Fill days
        for d in range(1, max_day + 1):
            val = getattr(row, str(d), "")
            ws.cell(i, 3 + d, val).alignment = center

    # Output to bytes
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
